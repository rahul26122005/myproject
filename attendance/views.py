from django.conf import settings
from django.shortcuts import render, redirect
import openpyxl
import os
from django.http import JsonResponse, HttpResponse, FileResponse
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from .models import Student
from .forms import MonthYearForm, UploadFileForm, UserRegisterForm, MyclassForm, LoginForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
import logging
from concurrent.futures import ThreadPoolExecutor
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import user_passes_test, login_required
from django.views.generic import TemplateView
from django.urls import reverse_lazy
from django.db import transaction
from django.contrib import messages

# Configure logging
logging.basicConfig(level=logging.DEBUG)

# Index View
class IndexView(LoginRequiredMixin, View):
    def get(self, request):

        return render(request, 'index.html')

# Group-specific View
@method_decorator(user_passes_test(lambda u: u.is_superuser or Group.objects.filter(name='YourGroupName', user=u).exists()), name='dispatch')
class GroupSpecificView(View):
    def get(self, request):
        return render(request, 'group_specific.html')

# Custom Login View
class CustomLoginView(View):
    form_class = LoginForm
    template_name = 'login.html'

    def get(self, request):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            user = authenticate(request, username=form.cleaned_data['username'], password=form.cleaned_data['password'])
            if user is not None:
                login(request, user)
                return redirect('home')
        return render(request, self.template_name, {'form': form})

# Custom Logout View
class CustomLogoutView(View):
    def get(self, request):
        logout(request)
        messages.success(request, ("You Are Successfully Logged Out !!...  "))
        return redirect('home')

# Registration View
class RegisterView(LoginRequiredMixin, View):
    def get(self, request):
        form = UserRegisterForm()
        return render(request, 'register.html', {'form': form})

    def post(self, request):
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            group, _ = Group.objects.get_or_create(name='YourGroupName')
            user.groups.add(group)
            raw_password = form.cleaned_data.get('password1')
            user = authenticate(username=user.username, password=raw_password)
            login(request, user)
            messages.success(request, ("You Are Successfully Logged In !!...  "))
            messages.success(request, ("wellcome {username} "))
            return redirect('home')
        return render(request, 'register.html', {'form': form})

# Upload Students View

class UploadStudentFileView(View):
    def get(self, request):
        form = UploadFileForm()
        return render(request, 'upload_students.html', {'form': form})

    def post(self, request):
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            missing_details = []

            with transaction.atomic():
                for row in sheet.iter_rows(min_row=2, values_only=True): # type: ignore
                    name, roll_number, student_class, section = row
                    if not all([name, roll_number, student_class, section]):
                        missing_details.append(row)
                    else:
                        Student.objects.create(
                            name=name,
                            roll_number=roll_number,
                            student_class=student_class,
                            section=section
                        )

            if missing_details:
                return render(request, 'upload_students.html', {
                    'form': form,
                    'missing_details': missing_details
                })

            return messages.success(request, ("You Are Successfully Upload the Students details  !!...  "))
        return render(request, 'upload_students.html', {'form': form})



# Student List View

class StudentListView(View):
    def get(self, request):
        file_path = os.path.join(settings.BASE_DIR, 'templates', 'student_template.xlsx')
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        students = []
        for row in sheet.iter_rows(min_row=2, values_only=True): # type: ignore
            students.append({
                'name': row[0],
                'roll_number': row[1],
                'department': row[2],
            })

        return render(request, 'attendance/student_list.html', {'students': students})

# Download Template View

def DownloadTemplateView(request):
    file_path = os.path.join(settings.BASE_DIR, 'templates', 'student_template.xlsx')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='student_template.xlsx')

class SelectClassView(View):
    def get(self, request):
        classes = Student.objects.values_list('student_class', flat=True).distinct()

        return HttpResponse(request, 'select_class.html', {'classes': classes})

# Attendance Form View
class View1(LoginRequiredMixin, View):
    def get(self, request):
        classes = Student.objects.values_list('student_class', flat=True).distinct()
        selected_class = request.GET.get('class')
        sections =  Student.objects.filter(student_class=selected_class).values_list('section', flat=True).distinct() if selected_class else []
        selected_section = request.GET.get('section')
        students =  Student.objects.filter(student_class=selected_class, section=selected_section) if selected_class and selected_section else [] # type: ignore
        form = MyclassForm()

        return render(request, 'Attendance.html', {
            'form': form,
            'students': students,
            'classes': classes,
            'selected_class': selected_class,
            'sections': sections,
            'selected_section': selected_section
        })

    def post(self, request):
        selected_class = request.POST.get('class')
        selected_section = request.POST.get('section')
        students =  Student.objects.filter(student_class=selected_class, section=selected_section)

        form_data = []
        for student in students:
            status = request.POST.get(f'status_{student.id}') # type: ignore
            if status:
                form_data.append({'student': student.id, 'status': status}) # type: ignore

        with transaction.atomic():
            for data in form_data:
                form = MyclassForm(data)
                if form.is_valid():
                    form.save()
                else:
                    logging.error("Form is not valid: %s", form.errors)
                    return JsonResponse({'error': 'Form is not valid', 'details': form.errors}, status=400)

        return JsonResponse({'message': 'Attendance marked successfully'})

# Report Generation View
class View2(View):
    template_name = 'generate_report.html'
    form_class = MonthYearForm

    def get(self, request):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            month = form.cleaned_data['month']
            year = form.cleaned_data['year']
            student_class = form.cleaned_data['student_class']
            section = form.cleaned_data['section']

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Attendance Record' # type: ignore

            # Setup Title and Header
            sheet['A1'] = 'ATTENDANCE RECORD' # type: ignore
            sheet['A1'].font = Font(size=14, bold=True) # type: ignore
            sheet['A1'].alignment = Alignment(horizontal='center') # type: ignore
            sheet.merge_cells('A1:H1') # type: ignore

            sheet['B2'] = 'Month:' # type: ignore
            sheet['C2'] = datetime(year, month, 1).strftime('%B') # type: ignore
            sheet['B3'] = 'Year:' # type: ignore
            sheet['C3'] = year # type: ignore

            sheet['B2'].alignment = Alignment(horizontal='right') # type: ignore
            sheet['B3'].alignment = Alignment(horizontal='right') # type: ignore

            headers = ['ST Name', 'Reg Number', 'Class', 'Section']
            for day in range(1, 32):
                try:
                    current_date = datetime(year, month, day)
                    headers.append(current_date.strftime('%a %d'))
                except ValueError:
                    break
            headers.append('Total Days Present')
            headers.append('Total Days Absents')

            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=5, column=col_num) # type: ignore
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                if col_num > 4:
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Optimized database query with prefetch_related
            students = Student.objects.filter(student_class=student_class, section=section).prefetch_related('myclass_set')

            def process_student(student):
                total_days_present = 0
                total_days_absent = 0
                student_row = [student.name, student.roll_number, student.student_class, student.section]
                for day in range(1, 32):
                    try:
                        current_date = datetime(year, month, day)
                        attendance = student.myclass_set.filter(date=current_date).first()
                        if attendance:
                            status = attendance.status
                            if status in ['present', 'od']:
                                total_days_present += 1
                            if status in ['absent']:
                                total_days_absent += 1
                        else:
                            status = ''
                        student_row.append(status)
                    except ValueError:
                        break
                student_row.append(total_days_present)
                student_row.append(total_days_absent)
                return student_row

            # Parallel processing using ThreadPoolExecutor
            with ThreadPoolExecutor(max_workers=10) as executor:
                student_rows = list(executor.map(process_student, students))
            
            for row_num, student_row in enumerate(student_rows, 6):
                for col_num, cell_value in enumerate(student_row, 1):
                    sheet.cell(row=row_num, column=col_num).value = cell_value # type: ignore

            for col_num in range(1, len(headers) + 1):
                sheet.column_dimensions[get_column_letter(col_num)].width = 15 # type: ignore

            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=attendance_{year}_{month}.xlsx'
            
            return response

        return render(request, self.template_name, {'form': form})
